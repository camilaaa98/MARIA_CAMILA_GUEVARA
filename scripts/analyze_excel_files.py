from openpyxl import load_workbook
import os

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'
CRISTIAN_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\cristian2.xlsx'

def analyze_files():
    """Analiza ambos archivos y muestra estructura"""
    
    print("=" * 80)
    print("ANÁLISIS DE ARCHIVOS EXCEL")
    print("=" * 80)
    
    # Leer Aprendices.xlsx
    print("\n📂 Analizando Aprendices.xlsx...")
    wb_aprendices = load_workbook(APRENDICES_PATH)
    ws_aprendices = wb_aprendices.active
    
    print(f"   Hoja activa: {ws_aprendices.title}")
    print(f"   Filas: {ws_aprendices.max_row}")
    print(f"   Columnas: {ws_aprendices.max_column}")
    
    # Leer encabezados
    headers_aprendices = []
    for col in range(1, ws_aprendices.max_column + 1):
        headers_aprendices.append(ws_aprendices.cell(1, col).value)
    print(f"   Encabezados: {headers_aprendices}")
    
    # Leer cristian2.xlsx
    print("\n📂 Analizando cristian2.xlsx...")
    wb_cristian = load_workbook(CRISTIAN_PATH)
    ws_cristian = wb_cristian.active
    
    print(f"   Hoja activa: {ws_cristian.title}")
    print(f"   Filas: {ws_cristian.max_row}")
    print(f"   Columnas: {ws_cristian.max_column}")
    
    # Leer encabezados
    headers_cristian = []
    for col in range(1, ws_cristian.max_column + 1):
        headers_cristian.append(ws_cristian.cell(1, col).value)
    print(f"   Encabezados: {headers_cristian}")
    
    # Mostrar primeras 3 filas de cada archivo
    print("\n" + "=" * 80)
    print("MUESTRA DE DATOS - Aprendices.xlsx (primeras 3 filas)")
    print("=" * 80)
    for row in range(1, min(4, ws_aprendices.max_row + 1)):
        row_data = []
        for col in range(1, ws_aprendices.max_column + 1):
            row_data.append(ws_aprendices.cell(row, col).value)
        print(f"Fila {row}: {row_data}")
    
    print("\n" + "=" * 80)
    print("MUESTRA DE DATOS - cristian2.xlsx (primeras 3 filas)")
    print("=" * 80)
    for row in range(1, min(4, ws_cristian.max_row + 1)):
        row_data = []
        for col in range(1, ws_cristian.max_column + 1):
            row_data.append(ws_cristian.cell(row, col).value)
        print(f"Fila {row}: {row_data}")
    
    # Identificar índices de columnas importantes
    print("\n" + "=" * 80)
    print("MAPEO DE COLUMNAS")
    print("=" * 80)
    
    # Buscar columnas en Aprendices
    ficha_idx_apr = None
    jornada_idx_apr = None
    nivel_idx_apr = None
    
    for i, header in enumerate(headers_aprendices):
        if header and 'ficha' in str(header).lower():
            ficha_idx_apr = i + 1
            print(f"Aprendices - Columna FICHA: '{header}' (índice {i+1})")
        if header and 'jornada' in str(header).lower():
            jornada_idx_apr = i + 1
            print(f"Aprendices - Columna JORNADA: '{header}' (índice {i+1})")
        if header and ('nivel' in str(header).lower() or 'formacion' in str(header).lower()):
            nivel_idx_apr = i + 1
            print(f"Aprendices - Columna NIVEL: '{header}' (índice {i+1})")
    
    # Buscar columnas en Cristian
    ficha_idx_cri = None
    jornada_idx_cri = None
    nivel_idx_cri = None
    instructor_idx_cri = None
    
    for i, header in enumerate(headers_cristian):
        if header and 'ficha' in str(header).lower():
            ficha_idx_cri = i + 1
            print(f"Cristian - Columna FICHA: '{header}' (índice {i+1})")
        if header and 'jornada' in str(header).lower():
            jornada_idx_cri = i + 1
            print(f"Cristian - Columna JORNADA: '{header}' (índice {i+1})")
        if header and ('nivel' in str(header).lower() or 'formacion' in str(header).lower()):
            nivel_idx_cri = i + 1
            print(f"Cristian - Columna NIVEL: '{header}' (índice {i+1})")
        if header and ('instructor' in str(header).lower() or 'lider' in str(header).lower()):
            instructor_idx_cri = i + 1
            print(f"Cristian - Columna INSTRUCTOR: '{header}' (índice {i+1})")
    
    # Extraer fichas únicas de cristian2
    if ficha_idx_cri:
        fichas_cristian = set()
        for row in range(2, ws_cristian.max_row + 1):
            ficha = ws_cristian.cell(row, ficha_idx_cri).value
            if ficha:
                fichas_cristian.add(ficha)
        
        print(f"\n📋 Fichas únicas en cristian2.xlsx ({len(fichas_cristian)}):")
        for ficha in sorted(fichas_cristian):
            print(f"   • {ficha}")
    
    wb_aprendices.close()
    wb_cristian.close()

if __name__ == "__main__":
    analyze_files()
