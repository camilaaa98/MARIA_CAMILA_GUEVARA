from openpyxl import load_workbook
import os

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'

def inspect_updated_excel():
    print("=" * 80)
    print("INSPECCIONANDO APRENDICES.XLSX (ACTUALIZADO)")
    print("=" * 80)
    
    try:
        wb = load_workbook(APRENDICES_PATH)
        if 'fichas' not in wb.sheetnames:
            print("❌ No se encontró hoja 'fichas'")
            return
            
        ws = wb['fichas']
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        print(f"Encabezados: {headers}")
        
        print("\nMuestra de datos (primeras 3 filas):")
        for row in range(2, 5):
            values = [ws.cell(row, c).value for c in range(1, ws.max_column + 1)]
            print(f"  Fila {row}: {values}")
            
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    inspect_updated_excel()
