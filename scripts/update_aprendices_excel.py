from openpyxl import load_workbook
import shutil
from datetime import datetime
import os

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'
CRISTIAN_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\cristian2.xlsx'
BACKUP_DIR = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\backups'

def update_excel_with_cristian_data():
    """Actualiza Aprendices.xlsx con datos de cristian2.xlsx"""
    
    print("=" * 80)
    print("ACTUALIZANDO APRENDICES.XLSX CON DATOS DE CRISTIAN2")
    print("=" * 80)
    
    # 1. CREAR BACKUP
    print("\n📦 Creando backup...")
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"Aprendices_backup_{timestamp}.xlsx")
    shutil.copy2(APRENDICES_PATH, backup_path)
    print(f"✓ Backup creado: {backup_path}")
    
    # 2. LEER DATOS DE CRISTIAN2
    print("\n📂 Leyendo cristian2.xlsx...")
    wb_cristian = load_workbook(CRISTIAN_PATH)
    ws_cristian = wb_cristian.active
    
    # Encontrar columnas en cristian2
    headers_cri = [ws_cristian.cell(1, col).value for col in range(1, ws_cristian.max_column + 1)]
    
    ficha_idx_cri = next((i+1 for i, h in enumerate(headers_cri) if h and 'ficha' in str(h).lower()), None)
    jornada_idx_cri = next((i+1 for i, h in enumerate(headers_cri) if h and 'jornada' in str(h).lower()), None)
    instructor_idx_cri = next((i+1 for i, h in enumerate(headers_cri) if h and ('instructor' in str(h).lower() or 'lider' in str(h).lower())), None)
    
    print(f"   Columnas: Ficha={ficha_idx_cri}, Jornada={jornada_idx_cri}, Instructor={instructor_idx_cri}")
    
    # Crear diccionario con datos de cristian2
    datos_cristian = {}
    for row in range(2, ws_cristian.max_row + 1):
        ficha = ws_cristian.cell(row, ficha_idx_cri).value if ficha_idx_cri else None
        if ficha:
            datos_cristian[ficha] = {
                'jornada': ws_cristian.cell(row, jornada_idx_cri).value if jornada_idx_cri else None,
                'instructor_lider': ws_cristian.cell(row, instructor_idx_cri).value if instructor_idx_cri else None
            }
    
    print(f"✓ Cargados datos de {len(datos_cristian)} fichas únicas")
    wb_cristian.close()
    
    # 3. ACTUALIZAR APRENDICES.XLSX
    print("\n📝 Actualizando Aprendices.xlsx...")
    wb_apr = load_workbook(APRENDICES_PATH)
    ws_apr = wb_apr.active
    
    # Encontrar columnas en Aprendices
    headers_apr = [ws_apr.cell(1, col).value for col in range(1, ws_apr.max_column + 1)]
    
    ficha_idx_apr = next((i+1 for i, h in enumerate(headers_apr) if h and 'ficha' in str(h).lower()), None)
    jornada_idx_apr = next((i+1 for i, h in enumerate(headers_apr) if h and 'jornada' in str(h).lower()), None)
    instructor_idx_apr = next((i+1 for i, h in enumerate(headers_apr) if h and ('instructor' in str(h).lower() and 'lider' in str(h).lower())), None)
    
    # Si no existe columna instructor_lider, agregarla
    if not instructor_idx_apr:
        instructor_idx_apr = ws_apr.max_column + 1
        ws_apr.cell(1, instructor_idx_apr).value = 'instructor_lider'
        print(f"✓ Columna 'instructor_lider' agregada en posición {instructor_idx_apr}")
    
    # Actualizar datos
    actualizados = 0
    jornadas_cambiadas = 0
    instructores_agregados = 0
    
    for row in range(2, ws_apr.max_row + 1):
        ficha = ws_apr.cell(row, ficha_idx_apr).value if ficha_idx_apr else None
        
        if ficha and ficha in datos_cristian:
            datos = datos_cristian[ficha]
            
            # Actualizar jornada
            jornada_actual = ws_apr.cell(row, jornada_idx_apr).value if jornada_idx_apr else None
            if jornada_actual != datos['jornada']:
                ws_apr.cell(row, jornada_idx_apr).value = datos['jornada']
                jornadas_cambiadas += 1
            
            # Agregar/actualizar instructor_lider
            instructor_actual = ws_apr.cell(row, instructor_idx_apr).value if instructor_idx_apr else None
            if instructor_actual != datos['instructor_lider']:
                ws_apr.cell(row, instructor_idx_apr).value = datos['instructor_lider']
                instructores_agregados += 1
            
            actualizados += 1
    
    # Guardar cambios
    wb_apr.save(APRENDICES_PATH)
    wb_apr.close()
    
    print("\n" + "=" * 80)
    print("✅ ACTUALIZACIÓN COMPLETADA")
    print("=" * 80)
    print(f"\n📊 Resumen:")
    print(f"   Fichas actualizadas: {actualizados} de {len(datos_cristian)}")
    print(f"   Jornadas modificadas: {jornadas_cambiadas}")
    print(f"   Instructores agregados/actualizados: {instructores_agregados}")
    print(f"   Fichas NO tocadas: {ws_apr.max_row - 1 - actualizados}")
    print(f"\n💾 Backup en: {backup_path}")
    print(f"✓ Archivo actualizado: {APRENDICES_PATH}")

if __name__ == "__main__":
    update_excel_with_cristian_data()
