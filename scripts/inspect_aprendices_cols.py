from openpyxl import load_workbook
import os

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'

def inspect_aprendices():
    print("=" * 80)
    print("INSPECCIONANDO APRENDICES.XLSX")
    print("=" * 80)
    
    try:
        wb = load_workbook(APRENDICES_PATH)
        ws = wb.active
        
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        print(f"Encabezados encontrados: {headers}")
        
        # Buscar columna 'programas_formacion' o similar
        prog_col = None
        for i, h in enumerate(headers):
            if h and 'programas_formacion' in str(h).lower():
                prog_col = h
                print(f"✓ Encontrada columna exacta: '{h}' en índice {i+1}")
            elif h and 'programa' in str(h).lower():
                print(f"⚠ Posible coincidencia: '{h}' en índice {i+1}")
                
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    inspect_aprendices()
