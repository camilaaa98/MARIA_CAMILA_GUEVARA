from openpyxl import load_workbook
import os

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'

def inspect_all_sheets():
    print("=" * 80)
    print("INSPECCIONANDO TODAS LAS HOJAS")
    print("=" * 80)
    
    wb = load_workbook(APRENDICES_PATH)
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        print(f"\nHoja: '{sheet}'")
        print(f"Columnas: {headers}")
        print(f"Filas: {ws.max_row}")

if __name__ == "__main__":
    inspect_all_sheets()
