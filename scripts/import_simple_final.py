import sqlite3
import os
import hashlib
from openpyxl import load_workbook

# Configuración
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'Asistnet.db')
APRENDICES_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'Aprendices.xlsx')
CRISTIAN_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'cristian2.xlsx')

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def clean_text(text):
    if not text:
        return None
    return str(text).strip().upper()

def import_simple_data():
    print("=" * 80)
    print("IMPORTACIÓN A ESQUEMA SIMPLIFICADO")
    print("=" * 80)
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Limpiar (aunque acabamos de crear, por seguridad)
    cursor.execute("DELETE FROM asistencias")
    cursor.execute("DELETE FROM aprendices")
    cursor.execute("DELETE FROM asignaciones_instructor_ficha")
    cursor.execute("DELETE FROM fichas")
    cursor.execute("DELETE FROM programas_formacion")
    cursor.execute("DELETE FROM instructores")
    cursor.execute("DELETE FROM usuarios WHERE rol != 'administrador'")
    conn.commit()

    # Cargar mapa de instructores desde cristian2
    print("\n📂 Cargando instructores auxiliares...")
    mapa_instructores = {}
    try:
        wb_cri = load_workbook(CRISTIAN_PATH)
        ws_cri = wb_cri.active
        headers = [ws_cri.cell(1, c).value for c in range(1, ws_cri.max_column + 1)]
        
        idx_fi = None
        idx_in = None
        for i, h in enumerate(headers):
            if h:
                if 'ficha' in str(h).lower(): idx_fi = i + 1
                elif 'instructor' in str(h).lower() or 'lider' in str(h).lower(): idx_in = i + 1
        
        if idx_fi and idx_in:
            for row in range(2, ws_cri.max_row + 1):
                f = ws_cri.cell(row, idx_fi).value
                i = clean_text(ws_cri.cell(row, idx_in).value)
                if f and i:
                    mapa_instructores[f] = i
    except: pass

    print("\n📂 Procesando Aprendices.xlsx...")
    try:
        wb = load_workbook(APRENDICES_PATH)
    except Exception as e:
        print(f"❌ Error: {e}")
        return

    # 1. PROGRAMAS
    if 'programas_formacion' in wb.sheetnames:
        print("\n1️⃣  Importando Programas...")
        ws = wb['programas_formacion']
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        
        idx_nom = None
        idx_niv = None
        for i, h in enumerate(headers):
            if h:
                if 'nombre' in str(h).lower(): idx_nom = i + 1
                elif 'nivel' in str(h).lower(): idx_niv = i + 1
        
        count = 0
        for row in range(2, ws.max_row + 1):
            nom = clean_text(ws.cell(row, idx_nom).value) if idx_nom else None
            niv = clean_text(ws.cell(row, idx_niv).value) if idx_niv else None
            
            if nom:
                try:
                    cursor.execute("INSERT OR IGNORE INTO programas_formacion (nombre_programa, nivel_formacion) VALUES (?, ?)", (nom, niv))
                    count += 1
                except: pass
        print(f"   ✓ {count} programas importados.")

    # 2. FICHAS
    if 'fichas' in wb.sheetnames:
        print("\n2️⃣  Importando Fichas...")
        ws = wb['fichas']
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        
        # Mapeo basado en imagen: Col 1 es ID (ignorar), Col 2 es Numero
        idx_num = 2 
        
        idx_prog = None
        idx_jor = None
        idx_inst = None
        
        for i, h in enumerate(headers):
            if h:
                h_l = str(h).lower()
                if 'programa' in h_l: idx_prog = i + 1
                elif 'jornada' in h_l: idx_jor = i + 1
                elif 'intructor' in h_l or 'lider' in h_l: idx_inst = i + 1
        
        count = 0
        instructores_nuevos = set()
        
        for row in range(2, ws.max_row + 1):
            num = ws.cell(row, idx_num).value
            prog = clean_text(ws.cell(row, idx_prog).value) if idx_prog else None
            jor = clean_text(ws.cell(row, idx_jor).value) if idx_jor else None
            
            inst = clean_text(ws.cell(row, idx_inst).value) if idx_inst else None
            if not inst and num in mapa_instructores:
                inst = mapa_instructores[num]
            
            if num:
                try:
                    cursor.execute('''
                        INSERT OR REPLACE INTO fichas (numero_ficha, nombre_programa, jornada, estado, instructor_lider)
                        VALUES (?, ?, ?, 'ACTIVO', ?)
                    ''', (num, prog, jor, inst))
                    count += 1
                    if inst: instructores_nuevos.add(inst)
                except Exception as e:
                    # print(f"Err ficha {num}: {e}")
                    pass
        print(f"   ✓ {count} fichas importadas.")
        
        # Crear Instructores
        print("\n3️⃣  Creando Instructores...")
        count_inst = 0
        for inst_name in instructores_nuevos:
            parts = inst_name.split()
            if len(parts) > 0:
                nombre = parts[0]
                apellido = ' '.join(parts[1:]) if len(parts) > 1 else 'Instructor'
                email = f"{nombre.lower()}.{parts[-1].lower() if len(parts)>1 else 'sena'}@sena.edu.co"
                
                cursor.execute("SELECT 1 FROM usuarios WHERE correo = ?", (email,))
                if not cursor.fetchone():
                    pwd = hashlib.sha256('123456'.encode()).hexdigest()
                    try:
                        cursor.execute("INSERT INTO usuarios (nombre, apellido, correo, password_hash, rol, estado) VALUES (?, ?, ?, ?, 'instructor', 'activo')", (nombre, apellido, email, pwd))
                        uid = cursor.lastrowid
                        cursor.execute("INSERT INTO instructores (id_usuario, nombres, apellidos, correo, estado) VALUES (?, ?, ?, ?, 'activo')", (uid, nombre, apellido, email))
                        count_inst += 1
                    except: pass
        print(f"   ✓ {count_inst} instructores creados.")

    # 3. APRENDICES
    if 'aprendices' in wb.sheetnames:
        print("\n4️⃣  Importando Aprendices...")
        ws = wb['aprendices']
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        
        # Buscar columnas
        idx_doc = None
        idx_nom = None
        idx_ape = None
        idx_ficha = None
        idx_tipo = None
        idx_cel = None
        idx_mail = None
        
        for i, h in enumerate(headers):
            if h:
                h_l = str(h).lower()
                if 'documento' in h_l: idx_doc = i + 1
                elif 'nombre' in h_l: idx_nom = i + 1
                elif 'apellido' in h_l: idx_ape = i + 1
                elif 'numero_ficha' in h_l: idx_ficha = i + 1 # Prioridad numero_ficha
                elif 'ficha' in h_l and 'id' not in h_l: idx_ficha = i + 1
                elif 'tipo' in h_l: idx_tipo = i + 1
                elif 'celular' in h_l: idx_cel = i + 1
                elif 'correo' in h_l: idx_mail = i + 1
        
        # Fallback para ficha si no se encontró
        if not idx_ficha:
             for i, h in enumerate(headers):
                if h and 'id_ficha' in str(h).lower(): idx_ficha = i + 1 # Asumimos que aqui id_ficha es el numero

        count = 0
        for row in range(2, ws.max_row + 1):
            doc = ws.cell(row, idx_doc).value if idx_doc else None
            if not doc: continue
            
            nom = clean_text(ws.cell(row, idx_nom).value)
            ape = clean_text(ws.cell(row, idx_ape).value)
            ficha = ws.cell(row, idx_ficha).value if idx_ficha else None
            tipo = clean_text(ws.cell(row, idx_tipo).value) or 'CC'
            cel = ws.cell(row, idx_cel).value if idx_cel else None
            mail = clean_text(ws.cell(row, idx_mail).value)
            
            try:
                cursor.execute('''
                    INSERT OR REPLACE INTO aprendices (documento, tipo_identificacion, nombre, apellido, correo, celular, numero_ficha, estado)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 'EN FORMACION')
                ''', (doc, tipo, nom, ape, mail, cel, ficha))
                count += 1
            except: pass
        print(f"   ✓ {count} aprendices importados.")

    conn.commit()
    conn.close()
    print("\n✅ IMPORTACIÓN SIMPLIFICADA COMPLETADA")

if __name__ == "__main__":
    import_simple_data()
