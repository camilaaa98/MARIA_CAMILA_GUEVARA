from openpyxl import load_workbook
import os

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'
CRISTIAN_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\cristian2.xlsx'

def final_fix():
    print("=" * 80)
    print("CORRECCIÓN FINAL DE EXCEL")
    print("=" * 80)

    # 1. Cargar datos de referencia (cristian2.xlsx)
    print("\n1. Cargando datos de cristian2.xlsx...")
    wb_ref = load_workbook(CRISTIAN_PATH)
    ws_ref = wb_ref.active
    
    # Índices fijos basados en análisis previo
    # Ficha=3, Jornada=5, Instructor=6
    ref_data = {}
    for row in range(2, ws_ref.max_row + 1):
        try:
            ficha_val = ws_ref.cell(row, 3).value
            if ficha_val:
                ficha_str = str(ficha_val).strip()
                ref_data[ficha_str] = {
                    'jornada': ws_ref.cell(row, 5).value,
                    'instructor': ws_ref.cell(row, 6).value
                }
        except Exception as e:
            print(f"   Error leyendo fila {row} de referencia: {e}")
            
    print(f"   ✓ {len(ref_data)} fichas cargadas en memoria.")
    wb_ref.close()

    # 2. Actualizar destino (Aprendices.xlsx)
    print("\n2. Actualizando Aprendices.xlsx...")
    wb_dest = load_workbook(APRENDICES_PATH)
    ws_dest = wb_dest.active
    
    # Verificar columnas
    headers = [ws_dest.cell(1, c).value for c in range(1, ws_dest.max_column + 1)]
    print(f"   Columnas actuales: {headers}")
    
    # Buscar índices
    idx_ficha = None
    idx_instructor = None
    idx_jornada = None
    
    for i, h in enumerate(headers):
        h_str = str(h).lower() if h else ""
        if 'ficha' in h_str:
            idx_ficha = i + 1
        elif 'instructor' in h_str and 'lider' in h_str:
            idx_instructor = i + 1
        elif 'jornada' in h_str:
            idx_jornada = i + 1
            
    print(f"   Índices detectados: Ficha={idx_ficha}, Instructor={idx_instructor}, Jornada={idx_jornada}")
    
    # Crear columnas si faltan
    if not idx_instructor:
        idx_instructor = ws_dest.max_column + 1
        ws_dest.cell(1, idx_instructor).value = "instructor_lider"
        print(f"   + Creando columna 'instructor_lider' en índice {idx_instructor}")
        
    if not idx_jornada:
        idx_jornada = ws_dest.max_column + 1
        ws_dest.cell(1, idx_jornada).value = "jornada"
        print(f"   + Creando columna 'jornada' en índice {idx_jornada}")

    # Aplicar actualizaciones
    count_updates = 0
    for row in range(2, ws_dest.max_row + 1):
        ficha_val = ws_dest.cell(row, idx_ficha).value
        if ficha_val:
            ficha_str = str(ficha_val).strip()
            
            if ficha_str in ref_data:
                data = ref_data[ficha_str]
                
                # Actualizar
                ws_dest.cell(row, idx_jornada).value = data['jornada']
                ws_dest.cell(row, idx_instructor).value = data['instructor']
                
                count_updates += 1
                if count_updates <= 3:
                    print(f"   ✓ Actualizado: Ficha {ficha_str} -> Jornada: {data['jornada']}, Inst: {data['instructor']}")

    print(f"\n   Total filas actualizadas: {count_updates}")
    
    try:
        wb_dest.save(APRENDICES_PATH)
        print("\n✅ Archivo guardado correctamente.")
    except PermissionError:
        print("\n❌ ERROR: No se pudo guardar. Cierra el archivo Excel e inténtalo de nuevo.")
    
    wb_dest.close()

if __name__ == "__main__":
    final_fix()
