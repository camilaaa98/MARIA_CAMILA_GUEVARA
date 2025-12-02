from openpyxl import load_workbook
import os

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'
OUTPUT_FILE = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\scripts\excel_structure.txt'

def inspect_all_sheets_to_file():
    print("Iniciando inspección...")
    wb = load_workbook(APRENDICES_PATH)
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write("ESTRUCTURA DE APRENDICES.XLSX\n")
        f.write("=============================\n\n")
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            f.write(f"HOJA: '{sheet}'\n")
            f.write(f"  Columnas: {headers}\n")
            f.write(f"  Filas: {ws.max_row}\n")
            f.write("-" * 40 + "\n")
            
    print(f"Estructura guardada en {OUTPUT_FILE}")

if __name__ == "__main__":
    inspect_all_sheets_to_file()
