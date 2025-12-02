import sqlite3
import os
import hashlib
from openpyxl import load_workbook

# Configuración
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'Asistnet.db')
APRENDICES_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'Aprendices.xlsx')

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def clean_text(text):
    if not text:
        return None
    return str(text).strip().upper()

def fix_instructors():
    print("=" * 80)
    print("CORRIGIENDO INSTRUCTORES DESDE APRENDICES.XLSX")
    print("=" * 80)
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        wb = load_workbook(APRENDICES_PATH)
        if 'fichas' not in wb.sheetnames:
            print("❌ No se encontró hoja 'fichas'")
            return
            
        ws = wb['fichas']
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        print(f"Encabezados: {headers}")
        
        # Buscar columna con typo 'intructor' o 'lider'
        idx_inst = None
        for i, h in enumerate(headers):
            if h and ('intructor' in str(h).lower() or 'lider' in str(h).lower()):
                idx_inst = i + 1
                print(f"✓ Columna instructor encontrada: '{h}' en índice {idx_inst}")
                break
        
        if not idx_inst:
            print("❌ No se encontró columna de instructor")
            return

        idx_num = None
        for i, h in enumerate(headers):
            if h and 'numero' in str(h).lower():
                idx_num = i + 1
                break

        print("\nActualizando Fichas y Creando Instructores...")
        count_updates = 0
        instructores_nuevos = set()
        
        for row in range(2, ws.max_row + 1):
            num = ws.cell(row, idx_num).value if idx_num else None
            inst = clean_text(ws.cell(row, idx_inst).value)
            
            if num and inst:
                # Actualizar ficha
                cursor.execute("UPDATE fichas SET instructor_lider = ? WHERE numero_ficha = ?", (inst, num))
                if cursor.rowcount > 0:
                    count_updates += 1
                    instructores_nuevos.add(inst)
        
        print(f"   ✓ {count_updates} fichas actualizadas con instructor.")
        
        # Crear usuarios
        count_inst = 0
        for inst_name in instructores_nuevos:
            parts = inst_name.split()
            if len(parts) > 0:
                nombre = parts[0]
                apellido = ' '.join(parts[1:]) if len(parts) > 1 else 'Instructor'
                email = f"{nombre.lower()}.{parts[-1].lower() if len(parts)>1 else 'sena'}@sena.edu.co"
                
                cursor.execute("SELECT 1 FROM usuarios WHERE correo = ?", (email,))
                if not cursor.fetchone():
                    pwd = hashlib.sha256('123456'.encode()).hexdigest()
                    try:
                        cursor.execute("INSERT INTO usuarios (nombre, apellido, correo, password_hash, rol, estado) VALUES (?, ?, ?, ?, 'instructor', 'activo')", (nombre, apellido, email, pwd))
                        uid = cursor.lastrowid
                        cursor.execute("INSERT INTO instructores (id_usuario, nombres, apellidos, correo, estado) VALUES (?, ?, ?, ?, 'activo')", (uid, nombre, apellido, email))
                        count_inst += 1
                    except Exception as e:
                        print(f"   Err crear usuario {email}: {e}")
                        
        print(f"   ✓ {count_inst} instructores creados.")
        
        conn.commit()
        
    except Exception as e:
        print(f"❌ Error: {e}")
    finally:
        conn.close()
        wb.close()

if __name__ == "__main__":
    fix_instructors()
