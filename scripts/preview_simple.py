from openpyxl import load_workbook
import os

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'
CRISTIAN_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\cristian2.xlsx'
OUTPUT_FILE = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\scripts\preview_cambios.txt'

def preview_changes_to_file():
    """Guarda el preview en un archivo de texto"""
    
    output_lines = []
    output_lines.append("=" * 80)
    output_lines.append("PREVIEW DE CAMBIOS (NO SE MODIFICARÁ NADA AÚN)")
    output_lines.append("=" * 80)
    
    # Leer cristian2.xlsx
    wb_cristian = load_workbook(CRISTIAN_PATH)
    ws_cristian = wb_cristian.active
    
    headers_cri = [ws_cristian.cell(1, col).value for col in range(1, ws_cristian.max_column + 1)]
    
    ficha_idx_cri = next((i+1 for i, h in enumerate(headers_cri) if h and 'ficha' in str(h).lower()), None)
    jornada_idx_cri = next((i+1 for i, h in enumerate(headers_cri) if h and 'jornada' in str(h).lower()), None)
    
    info_por_ficha = {}
    for row in range(2, ws_cristian.max_row + 1):
        ficha = ws_cristian.cell(row, ficha_idx_cri).value if ficha_idx_cri else None
        if ficha:
            info_por_ficha[ficha] = ws_cristian.cell(row, jornada_idx_cri).value if jornada_idx_cri else None
    
    wb_cristian.close()
    
    # Leer Aprendices.xlsx
    wb_apr = load_workbook(APRENDICES_PATH)
    ws_apr = wb_apr.active
    
    headers_apr = [ws_apr.cell(1, col).value for col in range(1, ws_apr.max_column + 1)]
    
    ficha_idx_apr = next((i+1 for i, h in enumerate(headers_apr) if h and 'ficha' in str(h).lower()), None)
    jornada_idx_apr = next((i+1 for i, h in enumerate(headers_apr) if h and 'jornada' in str(h).lower()), None)
    
    cambios = []
    sin_cambios = 0
    no_encontradas = 0
    
    for row in range(2, ws_apr.max_row + 1):
        ficha_apr = ws_apr.cell(row, ficha_idx_apr).value if ficha_idx_apr else None
        jornada_actual = ws_apr.cell(row, jornada_idx_apr).value if jornada_idx_apr else None
        
        if ficha_apr and ficha_apr in info_por_ficha:
            jornada_nueva = info_por_ficha[ficha_apr]
            if jornada_actual != jornada_nueva:
                cambios.append(f"Fila {row} | Ficha {ficha_apr} | '{jornada_actual}' → '{jornada_nueva}'")
            else:
                sin_cambios += 1
        elif ficha_apr:
            no_encontradas += 1
    
    output_lines.append(f"\n✓ Fichas en cristian2: {len(info_por_ficha)}")
    output_lines.append(f"\n🔄 CAMBIOS QUE SE APLICARÍAN: {len(cambios)}\n")
    
    for cambio in cambios:
        output_lines.append(cambio)
    
    output_lines.append(f"\n📊 RESUMEN:")
    output_lines.append(f"   Cambios: {len(cambios)}")
    output_lines.append(f"   Sin cambios: {sin_cambios}")
    output_lines.append(f"   No encontradas: {no_encontradas}")
    output_lines.append(f"   Total: {ws_apr.max_row - 1}")
    
    wb_apr.close()
    
    # Guardar en archivo
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_lines))
    
    print(f"Preview guardado en: {OUTPUT_FILE}")
    print(f"\nCambios detectados: {len(cambios)}")
    print(f"Primeros 5 cambios:")
    for cambio in cambios[:5]:
        print(f"  {cambio}")

if __name__ == "__main__":
    preview_changes_to_file()
