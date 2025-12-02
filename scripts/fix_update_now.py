from openpyxl import load_workbook
import shutil
from datetime import datetime
import os

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'
CRISTIAN_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\cristian2.xlsx'
BACKUP_DIR = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\backups'

def fix_and_update():
    """REALMENTE actualiza el archivo esta vez"""
    
    print("=" * 80)
    print("ACTUALIZACIÓN REAL - CORRIGIENDO AHORA")
    print("=" * 80)
    
    # 1. Leer cristian2
    print("\n1. Leyendo cristian2.xlsx...")
    wb_cri = load_workbook(CRISTIAN_PATH)
    ws_cri = wb_cri.active
    
    # Identificar columnas
    headers_cri = [ws_cri.cell(1, c).value for c in range(1, ws_cri.max_column + 1)]
    
    ficha_col_cri = None
    jornada_col_cri = None
    instructor_col_cri = None
    
    for i, h in enumerate(headers_cri):
        if h and 'ficha' in str(h).lower():
            ficha_col_cri = i + 1
        if h and 'jornada' in str(h).lower():
            jornada_col_cri = i + 1
        if h and ('instructor' in str(h).lower() or 'lider' in str(h).lower()):
            instructor_col_cri = i + 1
    
    print(f"   Ficha: col {ficha_col_cri}")
    print(f"   Jornada: col {jornada_col_cri}")
    print(f"   Instructor: col {instructor_col_cri}")
    
    # Extraer datos
    datos = {}
    for row in range(2, ws_cri.max_row + 1):
        ficha = ws_cri.cell(row, ficha_col_cri).value
        if ficha:
            datos[int(ficha)] = {
                'jornada': ws_cri.cell(row, jornada_col_cri).value,
                'instructor': ws_cri.cell(row, instructor_col_cri).value
            }
    
    print(f"   ✓ {len(datos)} fichas cargadas de cristian2")
    wb_cri.close()
    
    # 2. Actualizar Aprendices
    print("\n2. Actualizando Aprendices.xlsx...")
    wb_apr = load_workbook(APRENDICES_PATH)
    ws_apr = wb_apr.active
    
    # Encontrar columnas en Aprendices
    headers_apr = [ws_apr.cell(1, c).value for c in range(1, ws_apr.max_column + 1)]
    
    ficha_col_apr = None
    jornada_col_apr = None
    instructor_col_apr = None
    
    for i, h in enumerate(headers_apr):
        if h and 'ficha' in str(h).lower():
            ficha_col_apr = i + 1
            print(f"   Ficha: col {i+1} ({h})")
        if h and 'jornada' in str(h).lower():
            jornada_col_apr = i + 1
            print(f"   Jornada: col {i+1} ({h})")
        if h and 'instructor' in str(h).lower() and 'lider' in str(h).lower():
            instructor_col_apr = i + 1
            print(f"   Instructor líder: col {i+1} ({h})")
    
    if not instructor_col_apr:
        print("   ⚠ Columna instructor_lider NO encontrada, buscando alternativa...")
        for i, h in enumerate(headers_apr):
            print(f"     Col {i+1}: {h}")
    
    # Actualizar filas
    actualizados = 0
    jornadas_ok = 0
    instructores_ok = 0
    
    print(f"\n3. Aplicando cambios...")
    for row in range(2, ws_apr.max_row + 1):
        ficha_valor = ws_apr.cell(row, ficha_col_apr).value
        
        if ficha_valor:
            try:
                ficha = int(ficha_valor)
                if ficha in datos:
                    # Actualizar jornada
                    ws_apr.cell(row, jornada_col_apr).value = datos[ficha]['jornada']
                    jornadas_ok += 1
                    
                    # Actualizar instructor
                    ws_apr.cell(row, instructor_col_apr).value = datos[ficha]['instructor']
                    instructores_ok += 1
                    
                    actualizados += 1
                    
                    if actualizados <= 3:
                        print(f"   ✓ Fila {row}: Ficha {ficha} → Jornada: {datos[ficha]['jornada']}, Instructor: {datos[ficha]['instructor']}")
            except:
                pass
    
    # Guardar
    wb_apr.save(APRENDICES_PATH)
    print(f"\n4. Guardando archivo...")
    wb_apr.close()
    
    print("\n" + "=" * 80)
    print("✅ ACTUALIZACIÓN COMPLETADA")
    print("=" * 80)
    print(f"Fichas actualizadas: {actualizados}")
    print(f"Jornadas actualizadas: {jornadas_ok}")
    print(f"Instructores agregados: {instructores_ok}")

if __name__ == "__main__":
    fix_and_update()
