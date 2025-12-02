import sqlite3
import os
import hashlib
from openpyxl import load_workbook

# Configuración
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'Asistnet.db')
APRENDICES_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'Aprendices.xlsx')

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def clean_text(text):
    if not text:
        return None
    return str(text).strip().upper()

def find_col(headers, keywords):
    for i, h in enumerate(headers):
        if h:
            h_str = str(h).lower()
            for k in keywords:
                if k in h_str:
                    return i + 1
    return None

def import_all_from_aprendices():
    print("=" * 80)
    print("IMPORTACIÓN FINAL DESDE APRENDICES.XLSX")
    print("=" * 80)
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Limpiar datos previos (opcional, pero recomendado por el usuario)
    print("🧹 Limpiando tablas...")
    cursor.execute("DELETE FROM asistencias")
    cursor.execute("DELETE FROM aprendices")
    cursor.execute("DELETE FROM asignaciones_instructor_ficha")
    cursor.execute("DELETE FROM fichas")
    cursor.execute("DELETE FROM programas_formacion")
    cursor.execute("DELETE FROM instructores")
    cursor.execute("DELETE FROM usuarios WHERE rol != 'administrador'")
    conn.commit()
    
    try:
        wb = load_workbook(APRENDICES_PATH)
    except Exception as e:
        print(f"❌ Error leyendo archivo: {e}")
        return

    # 1. PROGRAMAS DE FORMACIÓN
    if 'programas_formacion' in wb.sheetnames:
        print("\n1️⃣  Importando Programas...")
        ws = wb['programas_formacion']
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        
        idx_nom = find_col(headers, ['nombre', 'programa'])
        idx_niv = find_col(headers, ['nivel'])
        idx_inst = find_col(headers, ['instructor', 'lider'])
        
        count = 0
        for row in range(2, ws.max_row + 1):
            nom = clean_text(ws.cell(row, idx_nom).value) if idx_nom else None
            niv = clean_text(ws.cell(row, idx_niv).value) if idx_niv else None
            inst = clean_text(ws.cell(row, idx_inst).value) if idx_inst else None
            
            if nom:
                try:
                    cursor.execute("INSERT OR IGNORE INTO programas_formacion (nombre_programa, nivel_formacion, instructor_lider) VALUES (?, ?, ?)", (nom, niv, inst))
                    count += 1
                except: pass
        print(f"   ✓ {count} programas importados.")

    # 2. FICHAS
    if 'fichas' in wb.sheetnames:
        print("\n2️⃣  Importando Fichas...")
        ws = wb['fichas']
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        
        idx_num = find_col(headers, ['numero', 'ficha'])
        idx_prog = find_col(headers, ['programa'])
        idx_jor = find_col(headers, ['jornada'])
        idx_inst = find_col(headers, ['instructor', 'lider'])
        
        count = 0
        instructores_nuevos = set()
        
        for row in range(2, ws.max_row + 1):
            num = ws.cell(row, idx_num).value if idx_num else None
            prog = clean_text(ws.cell(row, idx_prog).value) if idx_prog else None
            jor = clean_text(ws.cell(row, idx_jor).value) if idx_jor else None
            inst = clean_text(ws.cell(row, idx_inst).value) if idx_inst else None
            
            if num:
                try:
                    cursor.execute("INSERT OR REPLACE INTO fichas (numero_ficha, nombre_programa, jornada, estado, instructor_lider) VALUES (?, ?, ?, 'ACTIVO', ?)", (num, prog, jor, inst))
                    count += 1
                    if inst: instructores_nuevos.add(inst)
                except Exception as e:
                    print(f"Err ficha {num}: {e}")
        print(f"   ✓ {count} fichas importadas.")
        
        # Crear usuarios para instructores encontrados en fichas
        print("\n3️⃣  Creando Instructores...")
        count_inst = 0
        for inst_name in instructores_nuevos:
            parts = inst_name.split()
            if len(parts) > 0:
                nombre = parts[0]
                apellido = ' '.join(parts[1:]) if len(parts) > 1 else 'Instructor'
                email = f"{nombre.lower()}.{parts[-1].lower() if len(parts)>1 else 'sena'}@sena.edu.co"
                
                cursor.execute("SELECT 1 FROM usuarios WHERE correo = ?", (email,))
                if not cursor.fetchone():
                    pwd = hashlib.sha256('123456'.encode()).hexdigest()
                    cursor.execute("INSERT INTO usuarios (nombre, apellido, correo, password_hash, rol, estado) VALUES (?, ?, ?, ?, 'instructor', 'activo')", (nombre, apellido, email, pwd))
                    uid = cursor.lastrowid
                    cursor.execute("INSERT INTO instructores (id_usuario, nombres, apellidos, correo, estado) VALUES (?, ?, ?, ?, 'activo')", (uid, nombre, apellido, email))
                    count_inst += 1
        print(f"   ✓ {count_inst} instructores creados.")

    # 3. APRENDICES
    if 'aprendices' in wb.sheetnames:
        print("\n4️⃣  Importando Aprendices...")
        ws = wb['aprendices']
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        
        idx_doc = find_col(headers, ['documento'])
        idx_nom = find_col(headers, ['nombre'])
        idx_ape = find_col(headers, ['apellido'])
        idx_ficha = find_col(headers, ['ficha'])
        idx_tipo = find_col(headers, ['tipo'])
        idx_cel = find_col(headers, ['celular'])
        idx_mail = find_col(headers, ['correo'])
        
        count = 0
        for row in range(2, ws.max_row + 1):
            doc = ws.cell(row, idx_doc).value if idx_doc else None
            if not doc: continue
            
            nom = clean_text(ws.cell(row, idx_nom).value)
            ape = clean_text(ws.cell(row, idx_ape).value)
            ficha = ws.cell(row, idx_ficha).value if idx_ficha else None
            tipo = clean_text(ws.cell(row, idx_tipo).value) or 'CC'
            cel = ws.cell(row, idx_cel).value if idx_cel else None
            mail = clean_text(ws.cell(row, idx_mail).value)
            
            try:
                cursor.execute('''
                    INSERT OR REPLACE INTO aprendices (documento, tipo_identificacion, nombre, apellido, correo, celular, numero_ficha, estado)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 'EN FORMACION')
                ''', (doc, tipo, nom, ape, mail, cel, ficha))
                count += 1
            except: pass
        print(f"   ✓ {count} aprendices importados.")

    conn.commit()
    conn.close()
    print("\n✅ PROCESO FINALIZADO")

if __name__ == "__main__":
    import_all_from_aprendices()
