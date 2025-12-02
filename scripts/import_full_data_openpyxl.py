import sqlite3
import os
import hashlib
from openpyxl import load_workbook

# Configuración
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'Asistnet.db')
CRISTIAN_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'cristian2.xlsx')
APRENDICES_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'Aprendices.xlsx')

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def clean_text(text):
    if not text:
        return None
    return str(text).strip().upper()

def find_column_index(headers, keywords):
    for i, h in enumerate(headers):
        if h:
            h_str = str(h).lower()
            for k in keywords:
                if k in h_str:
                    return i + 1 # 1-based index for openpyxl
    return None

def import_data():
    print("=" * 80)
    print("INICIANDO IMPORTACIÓN MASIVA DE DATOS (OPENPYXL)")
    print("=" * 80)
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # ---------------------------------------------------------
    # 1. IMPORTAR PROGRAMAS Y FICHAS (desde cristian2.xlsx)
    # ---------------------------------------------------------
    print("\n📂 Leyendo cristian2.xlsx...")
    try:
        wb_cri = load_workbook(CRISTIAN_PATH)
        ws_cri = wb_cri.active
    except Exception as e:
        print(f"❌ Error leyendo cristian2.xlsx: {e}")
        return

    # Leer encabezados
    headers_cri = [ws_cri.cell(1, c).value for c in range(1, ws_cri.max_column + 1)]
    
    # Identificar columnas
    idx_ficha = find_column_index(headers_cri, ['ficha'])
    idx_jornada = find_column_index(headers_cri, ['jornada'])
    idx_nivel = find_column_index(headers_cri, ['nivel'])
    idx_prog = find_column_index(headers_cri, ['formacion']) # Cuidado con 'nivel de formacion'
    # Refinar busqueda de programa si confundió con nivel
    if idx_prog == idx_nivel:
        # Buscar otro que tenga formacion pero no nivel
        for i, h in enumerate(headers_cri):
            if h and 'formacion' in str(h).lower() and 'nivel' not in str(h).lower():
                idx_prog = i + 1
                break
                
    idx_inst = find_column_index(headers_cri, ['instructor', 'lider'])

    print(f"   Columnas: Ficha={idx_ficha}, Prog={idx_prog}, Nivel={idx_nivel}, Jornada={idx_jornada}, Inst={idx_inst}")

    # Estructuras para evitar duplicados
    programas_vistos = set()
    instructores_vistos = set()
    
    count_prog = 0
    count_inst = 0
    count_fichas = 0
    
    print("\n1️⃣  Procesando Programas, Instructores y Fichas...")
    
    for row in range(2, ws_cri.max_row + 1):
        # Leer valores
        ficha_val = ws_cri.cell(row, idx_ficha).value if idx_ficha else None
        prog_val = clean_text(ws_cri.cell(row, idx_prog).value) if idx_prog else None
        nivel_val = clean_text(ws_cri.cell(row, idx_nivel).value) if idx_nivel else None
        jornada_val = clean_text(ws_cri.cell(row, idx_jornada).value) if idx_jornada else None
        inst_val = clean_text(ws_cri.cell(row, idx_inst).value) if idx_inst else None
        
        if not ficha_val: continue
        
        # A. PROGRAMAS
        if prog_val and prog_val not in programas_vistos:
            try:
                cursor.execute('''
                    INSERT OR IGNORE INTO programas_formacion (nombre_programa, nivel_formacion, instructor_lider)
                    VALUES (?, ?, ?)
                ''', (prog_val, nivel_val, inst_val))
                programas_vistos.add(prog_val)
                count_prog += 1
            except Exception as e:
                pass # print(f"Err prog: {e}")

        # B. INSTRUCTORES (Usuarios)
        if inst_val and inst_val not in instructores_vistos:
            parts = inst_val.split()
            if len(parts) > 0:
                nombre = parts[0]
                apellido = ' '.join(parts[1:]) if len(parts) > 1 else 'Instructor'
                email = f"{nombre.lower()}.{parts[-1].lower() if len(parts)>1 else 'sena'}@sena.edu.co"
                
                # Verificar existencia
                cursor.execute("SELECT id_usuario FROM usuarios WHERE correo = ?", (email,))
                if not cursor.fetchone():
                    try:
                        pwd_hash = hashlib.sha256('123456'.encode()).hexdigest()
                        cursor.execute('''
                            INSERT INTO usuarios (nombre, apellido, correo, password_hash, rol, estado)
                            VALUES (?, ?, ?, ?, 'instructor', 'activo')
                        ''', (nombre, apellido, email, pwd_hash))
                        uid = cursor.lastrowid
                        
                        cursor.execute('''
                            INSERT INTO instructores (id_usuario, nombres, apellidos, correo, estado)
                            VALUES (?, ?, ?, ?, 'activo')
                        ''', (uid, nombre, apellido, email))
                        count_inst += 1
                    except:
                        pass # Email duplicado o error
            instructores_vistos.add(inst_val)

        # C. FICHAS
        try:
            cursor.execute('''
                INSERT OR REPLACE INTO fichas (numero_ficha, nombre_programa, jornada, estado, instructor_lider)
                VALUES (?, ?, ?, 'ACTIVO', ?)
            ''', (ficha_val, prog_val, jornada_val, inst_val))
            count_fichas += 1
        except Exception as e:
            print(f"   Err ficha {ficha_val}: {e}")

    print(f"   ✓ Programas: {count_prog}")
    print(f"   ✓ Instructores: {count_inst}")
    print(f"   ✓ Fichas: {count_fichas}")
    
    wb_cri.close()

    # ---------------------------------------------------------
    # 2. IMPORTAR APRENDICES (desde Aprendices.xlsx)
    # ---------------------------------------------------------
    print("\n📂 Leyendo Aprendices.xlsx...")
    try:
        wb_apr = load_workbook(APRENDICES_PATH)
        ws_apr = wb_apr.active
    except Exception as e:
        print(f"❌ Error leyendo Aprendices.xlsx: {e}")
        conn.close()
        return

    headers_apr = [ws_apr.cell(1, c).value for c in range(1, ws_apr.max_column + 1)]
    
    idx_doc = find_column_index(headers_apr, ['documento'])
    idx_nom = find_column_index(headers_apr, ['nombre'])
    idx_ape = find_column_index(headers_apr, ['apellido'])
    idx_ficha_apr = find_column_index(headers_apr, ['ficha'])
    idx_tipo = find_column_index(headers_apr, ['tipo'])
    idx_cel = find_column_index(headers_apr, ['celular'])
    idx_mail = find_column_index(headers_apr, ['correo'])

    print(f"   Columnas: Doc={idx_doc}, Nom={idx_nom}, Ficha={idx_ficha_apr}")

    count_apr = 0
    errores_apr = 0
    
    print("\n2️⃣  Importando Aprendices...")
    
    for row in range(2, ws_apr.max_row + 1):
        try:
            doc_val = ws_apr.cell(row, idx_doc).value if idx_doc else None
            if not doc_val: continue
            
            ficha_val = ws_apr.cell(row, idx_ficha_apr).value if idx_ficha_apr else None
            nom_val = clean_text(ws_apr.cell(row, idx_nom).value)
            ape_val = clean_text(ws_apr.cell(row, idx_ape).value)
            tipo_val = clean_text(ws_apr.cell(row, idx_tipo).value) or 'CC'
            mail_val = clean_text(ws_apr.cell(row, idx_mail).value)
            cel_val = ws_apr.cell(row, idx_cel).value if idx_cel else None
            
            # Asegurar ficha existe
            if ficha_val:
                cursor.execute("SELECT 1 FROM fichas WHERE numero_ficha = ?", (ficha_val,))
                if not cursor.fetchone():
                    # Crear ficha generica si no existe
                    cursor.execute("INSERT OR IGNORE INTO programas_formacion (nombre_programa) VALUES ('GENERICO')")
                    cursor.execute('''
                        INSERT OR IGNORE INTO fichas (numero_ficha, nombre_programa, jornada, estado)
                        VALUES (?, 'GENERICO', 'NO DEFINIDA', 'ACTIVO')
                    ''', (ficha_val,))
            
            cursor.execute('''
                INSERT OR REPLACE INTO aprendices (documento, tipo_identificacion, nombre, apellido, correo, celular, numero_ficha, estado)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'EN FORMACION')
            ''', (doc_val, tipo_val, nom_val, ape_val, mail_val, cel_val, ficha_val))
            count_apr += 1
            
        except Exception as e:
            errores_apr += 1
            # print(f"Err apr row {row}: {e}")

    print(f"   ✓ Aprendices insertados: {count_apr}")
    print(f"   ⚠ Errores/Omitidos: {errores_apr}")

    conn.commit()
    conn.close()
    
    print("\n" + "=" * 80)
    print("✅ IMPORTACIÓN COMPLETADA")
    print("=" * 80)

if __name__ == "__main__":
    import_data()
