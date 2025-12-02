import sqlite3
import os
from openpyxl import load_workbook

DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'Asistnet.db')
CRISTIAN_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'cristian2.xlsx')

def check_intersection():
    print("=" * 80)
    print("CHECK INTERSECTION")
    print("=" * 80)
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Fichas en BD
    cursor.execute("SELECT numero_ficha FROM fichas")
    db_fichas = set(r[0] for r in cursor.fetchall())
    print(f"Fichas en BD: {len(db_fichas)}")
    
    # Fichas en Excel
    wb = load_workbook(CRISTIAN_PATH)
    ws = wb.active
    
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx_ficha = None
    for i, h in enumerate(headers):
        if h and 'ficha' in str(h).lower():
            idx_ficha = i + 1
            break
            
    excel_fichas = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, idx_ficha).value
        if val:
            excel_fichas.add(val)
            
    print(f"Fichas en Excel: {len(excel_fichas)}")
    
    # Intersección
    intersection = db_fichas.intersection(excel_fichas)
    print(f"Intersección: {len(intersection)}")
    
    if len(intersection) == 0:
        print("¡NO HAY COINCIDENCIAS!")
        print(f"Ejemplo BD: {list(db_fichas)[:5]}")
        print(f"Ejemplo Excel: {list(excel_fichas)[:5]}")

    conn.close()
    wb.close()

if __name__ == "__main__":
    check_intersection()
