import sqlite3
import os
from openpyxl import load_workbook

DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'Asistnet.db')
CRISTIAN_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'cristian2.xlsx')

def debug_matching():
    print("=" * 80)
    print("DEBUG MATCHING FICHAS")
    print("=" * 80)
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Obtener fichas de BD
    cursor.execute("SELECT numero_ficha FROM fichas LIMIT 5")
    db_fichas = [r[0] for r in cursor.fetchall()]
    print(f"BD Fichas (sample): {db_fichas} (Type: {[type(x) for x in db_fichas]})")
    
    # Obtener fichas de Excel
    wb = load_workbook(CRISTIAN_PATH)
    ws = wb.active
    
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx_ficha = None
    for i, h in enumerate(headers):
        if h and 'ficha' in str(h).lower():
            idx_ficha = i + 1
            break
            
    excel_fichas = []
    for row in range(2, 7):
        val = ws.cell(row, idx_ficha).value
        excel_fichas.append(val)
        
    print(f"Excel Fichas (sample): {excel_fichas} (Type: {[type(x) for x in excel_fichas]})")
    
    conn.close()
    wb.close()

if __name__ == "__main__":
    debug_matching()
