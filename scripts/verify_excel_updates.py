from openpyxl import load_workbook

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'

def verify_updates():
    """Verifica que las actualizaciones se hayan aplicado correctamente"""
    
    print("=" * 80)
    print("VERIFICACIÓN DE ACTUALIZACIONES")
    print("=" * 80)
    
    wb = load_workbook(APRENDICES_PATH)
    ws = wb.active
    
    # Obtener encabezados
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    
    print(f"\n📋 Columnas en el archivo:")
    for i, header in enumerate(headers, 1):
        if header:
            print(f"   {i}. {header}")
    
    # Verificar si existe columna instructor_lider
    instructor_col = next((i+1 for i, h in enumerate(headers) if h and 'instructor' in str(h).lower() and 'lider' in str(h).lower()), None)
    
    if instructor_col:
        print(f"\n✓ Columna 'instructor_lider' encontrada en posición {instructor_col}")
        
        # Contar cuántos registros tienen instructor_lider
        con_instructor = 0
        sin_instructor = 0
        
        for row in range(2, ws.max_row + 1):
            valor = ws.cell(row, instructor_col).value
            if valor and str(valor).strip():
                con_instructor += 1
            else:
                sin_instructor += 1
        
        print(f"\n📊 Estadísticas de instructor_lider:")
        print(f"   Con instructor líder: {con_instructor}")
        print(f"   Sin instructor líder: {sin_instructor}")
        
        # Mostrar algunos ejemplos
        print(f"\n📝 Primeros 5 registros con instructor_lider:")
        count = 0
        for row in range(2, ws.max_row + 1):
            valor = ws.cell(row, instructor_col).value
            if valor and str(valor).strip():
                ficha_col = next((i+1 for i, h in enumerate(headers) if h and 'ficha' in str(h).lower()), None)
                ficha = ws.cell(row, ficha_col).value if ficha_col else "?"
                print(f"   Ficha {ficha}: {valor}")
                count += 1
                if count >= 5:
                    break
    else:
        print("\n✗ Columna 'instructor_lider' NO encontrada")
    
    wb.close()
    
    print("\n" + "=" * 80)
    print("VERIFICACIÓN COMPLETADA")
    print("=" * 80)

if __name__ == "__main__":
    verify_updates()
