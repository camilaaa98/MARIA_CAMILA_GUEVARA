import sqlite3
import os
import hashlib
from openpyxl import load_workbook

# Configuración
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'Asistnet.db')
APRENDICES_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'Aprendices.xlsx')
CRISTIAN_PATH = os.path.join(os.path.dirname(__file__), '..', 'database', 'cristian2.xlsx')

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def clean_text(text):
    if not text:
        return None
    return str(text).strip().upper()

def import_corrected_data():
    print("=" * 80)
    print("IMPORTACIÓN CORREGIDA Y FINAL (CON MAPEO EXACTO)")
    print("=" * 80)
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 1. Limpiar BD
    print("🧹 Limpiando base de datos...")
    cursor.execute("DELETE FROM asistencias")
    cursor.execute("DELETE FROM aprendices")
    cursor.execute("DELETE FROM asignaciones_instructor_ficha")
    cursor.execute("DELETE FROM fichas")
    cursor.execute("DELETE FROM programas_formacion")
    cursor.execute("DELETE FROM instructores")
    cursor.execute("DELETE FROM usuarios WHERE rol != 'administrador'")
    conn.commit()

    # 2. Cargar diccionarios auxiliares desde cristian2 (para instructores faltantes)
    print("\n📂 Cargando datos auxiliares de cristian2.xlsx...")
    mapa_instructores = {} # ficha -> instructor
    try:
        wb_cri = load_workbook(CRISTIAN_PATH)
        ws_cri = wb_cri.active
        headers_cri = [ws_cri.cell(1, c).value for c in range(1, ws_cri.max_column + 1)]
        
        idx_fi_cri = None
        idx_in_cri = None
        for i, h in enumerate(headers_cri):
            if h:
                if 'ficha' in str(h).lower(): idx_fi_cri = i + 1
                elif 'instructor' in str(h).lower() or 'lider' in str(h).lower(): idx_in_cri = i + 1
        
        if idx_fi_cri and idx_in_cri:
            for row in range(2, ws_cri.max_row + 1):
                f = ws_cri.cell(row, idx_fi_cri).value
                i = clean_text(ws_cri.cell(row, idx_in_cri).value)
                if f and i:
                    mapa_instructores[f] = i
        print(f"   ✓ {len(mapa_instructores)} instructores auxiliares cargados.")
    except Exception as e:
        print(f"   ⚠ No se pudo cargar cristian2: {e}")

    # 3. Importar desde Aprendices.xlsx
    print("\n📂 Procesando Aprendices.xlsx...")
    try:
        wb = load_workbook(APRENDICES_PATH)
    except Exception as e:
        print(f"❌ Error fatal leyendo archivo: {e}")
        return

    # A. PROGRAMAS
    if 'programas_formacion' in wb.sheetnames:
        print("\n1️⃣  Importando Programas...")
        ws = wb['programas_formacion']
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        
        # Buscar indices exactos
        idx_nom = None
        idx_niv = None
        for i, h in enumerate(headers):
            if h:
                h_l = str(h).lower()
                if 'nombre' in h_l and 'programa' in h_l: idx_nom = i + 1
                elif 'nivel' in h_l: idx_niv = i + 1
        
        count = 0
        for row in range(2, ws.max_row + 1):
            nom = clean_text(ws.cell(row, idx_nom).value) if idx_nom else None
            niv = clean_text(ws.cell(row, idx_niv).value) if idx_niv else None
            
            if nom:
                try:
                    cursor.execute("INSERT OR IGNORE INTO programas_formacion (nombre_programa, nivel_formacion) VALUES (?, ?)", (nom, niv))
                    count += 1
                except: pass
        print(f"   ✓ {count} programas importados.")

    # B. FICHAS (Corrección crítica de columnas)
    if 'fichas' in wb.sheetnames:
        print("\n2️⃣  Importando Fichas (Mapeo Corregido)...")
        ws = wb['fichas']
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        print(f"   Encabezados Fichas: {headers}")
        
        # Mapeo explicito basado en la imagen del usuario
        # Col 1: id_ficha (consecutivo)
        # Col 2: numero_ficha (real)
        idx_id = 1
        idx_num = 2
        
        # Buscar las otras columnas dinamicamente
        idx_prog = None
        idx_jor = None
        idx_inst = None
        
        for i, h in enumerate(headers):
            if h:
                h_l = str(h).lower()
                if 'programa' in h_l: idx_prog = i + 1
                elif 'jornada' in h_l: idx_jor = i + 1
                elif 'intructor' in h_l or 'lider' in h_l: idx_inst = i + 1
        
        count = 0
        instructores_nuevos = set()
        
        for row in range(2, ws.max_row + 1):
            id_val = ws.cell(row, idx_id).value # Col A
            num_val = ws.cell(row, idx_num).value # Col B
            
            prog = clean_text(ws.cell(row, idx_prog).value) if idx_prog else None
            jor = clean_text(ws.cell(row, idx_jor).value) if idx_jor else None
            
            # Intentar obtener instructor del excel, si no, del mapa auxiliar
            inst = clean_text(ws.cell(row, idx_inst).value) if idx_inst else None
            if not inst and num_val in mapa_instructores:
                inst = mapa_instructores[num_val]
            
            if num_val:
                try:
                    # Insertamos id_ficha explícitamente y numero_ficha explícitamente
                    cursor.execute('''
                        INSERT OR REPLACE INTO fichas (id_ficha, numero_ficha, nombre_programa, jornada, estado, instructor_lider)
                        VALUES (?, ?, ?, ?, 'ACTIVO', ?)
                    ''', (id_val, num_val, prog, jor, inst))
                    count += 1
                    if inst: instructores_nuevos.add(inst)
                except Exception as e:
                    print(f"   ⚠ Error ficha {num_val}: {e}")
        print(f"   ✓ {count} fichas importadas correctamente.")
        
        # Crear Instructores
        print("\n3️⃣  Creando Usuarios Instructores...")
        count_inst = 0
        for inst_name in instructores_nuevos:
            parts = inst_name.split()
            if len(parts) > 0:
                nombre = parts[0]
                apellido = ' '.join(parts[1:]) if len(parts) > 1 else 'Instructor'
                email = f"{nombre.lower()}.{parts[-1].lower() if len(parts)>1 else 'sena'}@sena.edu.co"
                
                cursor.execute("SELECT 1 FROM usuarios WHERE correo = ?", (email,))
                if not cursor.fetchone():
                    pwd = hashlib.sha256('123456'.encode()).hexdigest()
                    try:
                        cursor.execute("INSERT INTO usuarios (nombre, apellido, correo, password_hash, rol, estado) VALUES (?, ?, ?, ?, 'instructor', 'activo')", (nombre, apellido, email, pwd))
                        uid = cursor.lastrowid
                        cursor.execute("INSERT INTO instructores (id_usuario, nombres, apellidos, correo, estado) VALUES (?, ?, ?, ?, 'activo')", (uid, nombre, apellido, email))
                        count_inst += 1
                    except: pass
        print(f"   ✓ {count_inst} instructores creados.")

    # C. APRENDICES
    if 'aprendices' in wb.sheetnames:
        print("\n4️⃣  Importando Aprendices...")
        ws = wb['aprendices']
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        
        idx_doc = None
        idx_nom = None
        idx_ape = None
        idx_ficha = None
        idx_tipo = None
        idx_cel = None
        idx_mail = None
        
        for i, h in enumerate(headers):
            if h:
                h_l = str(h).lower()
                if 'documento' in h_l: idx_doc = i + 1
                elif 'nombre' in h_l: idx_nom = i + 1
                elif 'apellido' in h_l: idx_ape = i + 1
                elif 'ficha' in h_l and 'id' not in h_l: idx_ficha = i + 1 # Evitar id_ficha si existe columna ficha
                elif 'tipo' in h_l: idx_tipo = i + 1
                elif 'celular' in h_l: idx_cel = i + 1
                elif 'correo' in h_l: idx_mail = i + 1
        
        # Si no encontró 'ficha', buscar 'id_ficha' como fallback pero verificar contenido
        if not idx_ficha:
             for i, h in enumerate(headers):
                if h and 'id_ficha' in str(h).lower(): idx_ficha = i + 1

        count = 0
        for row in range(2, ws.max_row + 1):
            doc = ws.cell(row, idx_doc).value if idx_doc else None
            if not doc: continue
            
            nom = clean_text(ws.cell(row, idx_nom).value)
            ape = clean_text(ws.cell(row, idx_ape).value)
            ficha = ws.cell(row, idx_ficha).value if idx_ficha else None
            tipo = clean_text(ws.cell(row, idx_tipo).value) or 'CC'
            cel = ws.cell(row, idx_cel).value if idx_cel else None
            mail = clean_text(ws.cell(row, idx_mail).value)
            
            try:
                cursor.execute('''
                    INSERT OR REPLACE INTO aprendices (documento, tipo_identificacion, nombre, apellido, correo, celular, numero_ficha, estado)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 'EN FORMACION')
                ''', (doc, tipo, nom, ape, mail, cel, ficha))
                count += 1
            except Exception as e:
                # print(f"Err apr: {e}")
                pass
        print(f"   ✓ {count} aprendices importados.")

    conn.commit()
    conn.close()
    print("\n✅ PROCESO FINALIZADO CORRECTAMENTE")

if __name__ == "__main__":
    import_corrected_data()
