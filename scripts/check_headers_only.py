from openpyxl import load_workbook
import os

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'

def check_headers():
    wb = load_workbook(APRENDICES_PATH)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"Columnas en Aprendices.xlsx: {headers}")
    
    has_prog = any('programa' in str(h).lower() or 'formacion' in str(h).lower() for h in headers if h)
    print(f"¿Tiene información de Programa? {'SÍ' if has_prog else 'NO'}")

if __name__ == "__main__":
    check_headers()
