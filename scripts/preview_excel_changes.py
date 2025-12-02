from openpyxl import load_workbook
import os
from datetime import datetime

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'
CRISTIAN_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\cristian2.xlsx'

def preview_changes():
    """Muestra qué cambios se harían SIN modificar archivos"""
    
    print("=" * 80)
    print("PREVIEW DE CAMBIOS (NO SE MODIFICARÁ NADA AÚN)")
    print("=" * 80)
    
    # Leer cristian2.xlsx y crear diccionario de datos por ficha
    wb_cristian = load_workbook(CRISTIAN_PATH)
    ws_cristian = wb_cristian.active
    
    # Encontrar índices de columnas en cristian2
    headers_cri = []
    for col in range(1, ws_cristian.max_column + 1):
        headers_cri.append(ws_cristian.cell(1, col).value)
    
    ficha_idx_cri = None
    jornada_idx_cri = None
    nivel_idx_cri = None
    formacion_idx_cri = None
    instructor_idx_cri = None
    
    for i, header in enumerate(headers_cri):
        if header and 'ficha' in str(header).lower():
            ficha_idx_cri = i + 1
        if header and 'jornada' in str(header).lower():
            jornada_idx_cri = i + 1
        if header and 'nivel' in str(header).lower():
            nivel_idx_cri = i + 1
        if header and 'formacion' in str(header).lower() and 'nivel' not in str(header).lower():
            formacion_idx_cri = i + 1
        if header and ('instructor' in str(header).lower() or 'lider' in str(header).lower()):
            instructor_idx_cri = i + 1
    
    print(f"\n📋 Columnas identificadas en cristian2.xlsx:")
    print(f"   Ficha: columna {ficha_idx_cri}")
    print(f"   Jornada: columna {jornada_idx_cri}")
    print(f"   Nivel: columna {nivel_idx_cri}")
    print(f"   Formación: columna {formacion_idx_cri}")
    print(f"   Instructor líder: columna {instructor_idx_cri}")
    
    # Crear diccionario con info de cristian2 por ficha
    info_por_ficha = {}
    for row in range(2, ws_cristian.max_row + 1):
        ficha = ws_cristian.cell(row, ficha_idx_cri).value if ficha_idx_cri else None
        if ficha:
            info_por_ficha[ficha] = {
                'jornada': ws_cristian.cell(row, jornada_idx_cri).value if jornada_idx_cri else None,
                'nivel': ws_cristian.cell(row, nivel_idx_cri).value if nivel_idx_cri else None,
                'formacion': ws_cristian.cell(row, formacion_idx_cri).value if formacion_idx_cri else None,
                'instructor_lider': ws_cristian.cell(row, instructor_idx_cri).value if instructor_idx_cri else None
            }
    
    print(f"\n✓ Datos extraídos de {len(info_por_ficha)} fichas en cristian2.xlsx")
    
    wb_cristian.close()
    
    # Leer Aprendices.xlsx
    wb_apr = load_workbook(APRENDICES_PATH)
    ws_apr = wb_apr.active
    
    # Encontrar índices de columnas en Aprendices
    headers_apr = []
    for col in range(1, ws_apr.max_column + 1):
        headers_apr.append(ws_apr.cell(1, col).value)
    
    ficha_idx_apr = None
    jornada_idx_apr = None
    
    for i, header in enumerate(headers_apr):
        if header and 'ficha' in str(header).lower():
            ficha_idx_apr = i + 1
        if header and 'jornada' in str(header).lower():
            jornada_idx_apr = i + 1
    
    print(f"\n📋 Columnas identificadas en Aprendices.xlsx:")
    print(f"   Ficha: columna {ficha_idx_apr}")
    print(f"   Jornada: columna {jornada_idx_apr}")
    
    # Comparar y mostrar cambios
    print("\n" + "=" * 80)
    print("CAMBIOS QUE SE APLICARÍAN")
    print("=" * 80)
    
    cambios = []
    sin_cambios = 0
    no_encontradas = 0
    
    for row in range(2, ws_apr.max_row + 1):
        ficha_apr = ws_apr.cell(row, ficha_idx_apr).value if ficha_idx_apr else None
        jornada_actual = ws_apr.cell(row, jornada_idx_apr).value if jornada_idx_apr else None
        
        if ficha_apr and ficha_apr in info_por_ficha:
            info_cristian = info_por_ficha[ficha_apr]
            jornada_nueva = info_cristian['jornada']
            
            if jornada_actual != jornada_nueva:
                cambios.append({
                    'fila': row,
                    'ficha': ficha_apr,
                    'jornada_actual': jornada_actual,
                    'jornada_nueva': jornada_nueva,
                    'nivel': info_cristian['nivel'],
                    'formacion': info_cristian['formacion'],
                    'instructor': info_cristian['instructor_lider']
                })
            else:
                sin_cambios += 1
        elif ficha_apr:
            no_encontradas += 1
    
    if cambios:
        print(f"\n🔄 Se cambiarían {len(cambios)} registros:\n")
        for i, cambio in enumerate(cambios[:10], 1):  # Mostrar primeros 10
            print(f"{i}. Fila {cambio['fila']} - Ficha {cambio['ficha']}")
            print(f"   Jornada: '{cambio['jornada_actual']}' → '{cambio['jornada_nueva']}'")
            print(f"   Info adicional: {cambio['nivel']} | {cambio['formacion']} | {cambio['instructor']}")
            print()
        
        if len(cambios) > 10:
            print(f"   ... y {len(cambios) - 10} cambios más")
    else:
        print("\n✓ No hay cambios necesarios - todos los datos coinciden")
    
    print(f"\n📊 RESUMEN:")
    print(f"   Registros que SE CAMBIARÍAN: {len(cambios)}")
    print(f"   Registros SIN cambios: {sin_cambios}")
    print(f"   Fichas NO encontradas en cristian2 (se mantienen igual): {no_encontradas}")
    print(f"   Total registros en Aprendices.xlsx: {ws_apr.max_row - 1}")
    
    wb_apr.close()
    
    print("\n" + "=" * 80)
    print("⚠️  IMPORTANTE: Este es solo un PREVIEW")
    print("=" * 80)
    print("NO se ha modificado ningún archivo.")
    print("Si apruebas estos cambios, se creará un backup automático primero.")

if __name__ == "__main__":
    preview_changes()
