from openpyxl import load_workbook
import os

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'

def inspect_fichas_values():
    print("=" * 80)
    print("INSPECCIONANDO VALORES DE FICHAS")
    print("=" * 80)
    
    wb = load_workbook(APRENDICES_PATH)
    ws = wb['fichas']
    
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"Encabezados: {headers}")
    
    idx_inst = None
    for i, h in enumerate(headers):
        if h and 'intructor' in str(h).lower():
            idx_inst = i + 1
            break
            
    if idx_inst:
        print(f"Columna instructor en índice: {idx_inst}")
        print("Primeros 10 valores:")
        for row in range(2, 12):
            val = ws.cell(row, idx_inst).value
            print(f"  Fila {row}: '{val}'")
    else:
        print("No se encontró columna instructor")

if __name__ == "__main__":
    inspect_fichas_values()
