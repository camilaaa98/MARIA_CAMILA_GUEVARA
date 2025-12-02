from openpyxl import load_workbook
import os

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'
CRISTIAN_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\cristian2.xlsx'

def clean_text(text):
    if not text:
        return None
    return str(text).strip().upper()

def compare_programs():
    print("=" * 80)
    print("COMPARACIÓN DE PROGRAMAS DE FORMACIÓN")
    print("=" * 80)
    
    # 1. Leer Programas de Aprendices.xlsx
    print("📂 Leyendo Aprendices.xlsx...")
    try:
        wb_apr = load_workbook(APRENDICES_PATH)
        if 'programas_formacion' not in wb_apr.sheetnames:
            print("❌ No se encontró la hoja 'programas_formacion' en Aprendices.xlsx")
            return
            
        ws_apr = wb_apr['programas_formacion']
        headers_apr = [ws_apr.cell(1, c).value for c in range(1, ws_apr.max_column + 1)]
        
        # Buscar columna nombre_programa
        idx_nom_apr = None
        for i, h in enumerate(headers_apr):
            if h and 'nombre' in str(h).lower() and 'programa' in str(h).lower():
                idx_nom_apr = i + 1
                break
        
        programas_aprendices = set()
        if idx_nom_apr:
            for row in range(2, ws_apr.max_row + 1):
                val = clean_text(ws_apr.cell(row, idx_nom_apr).value)
                if val:
                    programas_aprendices.add(val)
        
        print(f"   ✓ {len(programas_aprendices)} programas encontrados en Aprendices.xlsx")
        
    except Exception as e:
        print(f"❌ Error leyendo Aprendices.xlsx: {e}")
        return

    # 2. Leer Programas de cristian2.xlsx
    print("\n📂 Leyendo cristian2.xlsx...")
    try:
        wb_cri = load_workbook(CRISTIAN_PATH)
        ws_cri = wb_cri.active
        headers_cri = [ws_cri.cell(1, c).value for c in range(1, ws_cri.max_column + 1)]
        
        # Buscar columnas
        idx_prog_cri = None
        idx_niv_cri = None
        
        for i, h in enumerate(headers_cri):
            if h:
                h_str = str(h).lower()
                if 'formacion' in h_str and 'nivel' not in h_str:
                    idx_prog_cri = i + 1
                elif 'nivel' in h_str:
                    idx_niv_cri = i + 1
        
        programas_cristian = {} # Nombre -> Nivel
        
        if idx_prog_cri:
            for row in range(2, ws_cri.max_row + 1):
                prog = clean_text(ws_cri.cell(row, idx_prog_cri).value)
                niv = clean_text(ws_cri.cell(row, idx_niv_cri).value) if idx_niv_cri else "NO DEFINIDO"
                
                if prog:
                    programas_cristian[prog] = niv
        
        print(f"   ✓ {len(programas_cristian)} programas encontrados en cristian2.xlsx")
        
    except Exception as e:
        print(f"❌ Error leyendo cristian2.xlsx: {e}")
        return

    # 3. Comparar y guardar en archivo
    output_file = 'reporte_programas.txt'
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("RESULTADOS: PROGRAMAS FALTANTES EN APRENDICES.XLSX\n")
        f.write("=" * 80 + "\n\n")
        
        faltantes = []
        for prog, niv in programas_cristian.items():
            if prog not in programas_aprendices:
                faltantes.append((prog, niv))
                
        if not faltantes:
            f.write("¡No faltan programas! Todos los de cristian2 están en Aprendices.xlsx\n")
        else:
            f.write(f"Se encontraron {len(faltantes)} programas nuevos/faltantes:\n\n")
            f.write(f"{'PROGRAMA':<50} | {'NIVEL':<20}\n")
            f.write("-" * 75 + "\n")
            for prog, niv in sorted(faltantes):
                f.write(f"{prog:<50} | {niv:<20}\n")
    
    print(f"Reporte guardado en {output_file}")

if __name__ == "__main__":
    compare_programs()
