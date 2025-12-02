from openpyxl import load_workbook

APRENDICES_PATH = r'C:\wamp64\www\YanguasEjercicios\mockups-asist-net\database\Aprendices.xlsx'

def check_real_status():
    """Verificación REAL del estado del archivo"""
    
    wb = load_workbook(APRENDICES_PATH)
    ws = wb.active
    
    print("=" * 80)
    print("ESTADO REAL DEL ARCHIVO")
    print("=" * 80)
    
    # Leer TODOS los encabezados
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(ws.cell(1, col).value)
    
    print(f"\nColumnas actuales ({len(headers)}):")
    for i, h in enumerate(headers, 1):
        print(f"  {i}. {h}")
    
    # Buscar instructor_lider
    tiene_instructor = any('instructor' in str(h).lower() and 'lider' in str(h).lower() for h in headers if h)
    
    print(f"\n¿Tiene columna instructor_lider? {'SÍ' if tiene_instructor else 'NO'}")
    
    # Verificar primeras 10 filas completas
    print(f"\nPrimeras 3 filas de datos:")
    for row in range(2, 5):
        print(f"\nFila {row}:")
        for col in range(1, min(ws.max_column + 1, 15)):
            valor = ws.cell(row, col).value
            print(f"  {headers[col-1]}: {valor}")
    
    wb.close()

if __name__ == "__main__":
    check_real_status()
